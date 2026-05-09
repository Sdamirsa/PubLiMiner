"""Export a PubLiMiner extraction run to a self-contained JSON snapshot.

Reads papers.parquet and extractions.db in read-only mode — safe to run
while the extract step is still in progress.

Usage:
    uv run python scripts/per-project/export_run_json.py --config projects/cardiac_mri.yaml
    uv run python scripts/per-project/export_run_json.py --config projects/cardiac_mri.yaml --run-id 20260506T083742Z
    uv run python scripts/per-project/export_run_json.py --config projects/cardiac_ct.yaml

Output:
    {output_dir}/snapshot_{schema_name}_{run_id}.json
"""

from __future__ import annotations

import argparse
import csv
import json
import sqlite3
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import polars as pl
import yaml


# ── helpers ────────────────────────────────────────────────────────────────────

def _safe_json(s: str | None) -> dict | list | None:
    if not s:
        return None
    try:
        return json.loads(s)
    except Exception:
        return None


def _journal_title(journal_json: str | None) -> str | None:
    j = _safe_json(journal_json)
    if isinstance(j, dict):
        return j.get("title") or j.get("journal_title") or j.get("iso_abbrev")
    if isinstance(j, str):
        return j
    return None


def _load_yaml(config_path: Path) -> dict:
    with config_path.open(encoding="utf-8") as f:
        return yaml.safe_load(f)


def _pick_run_id(db_path: Path, schema_name: str, requested: str | None) -> str:
    """Return the run_id to snapshot; defaults to the most recent one in the DB."""
    if requested:
        return requested
    con = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
    try:
        row = con.execute(
            "SELECT run_id FROM extractions WHERE schema_name = ? "
            "ORDER BY created_at DESC LIMIT 1",
            (schema_name,),
        ).fetchone()
    finally:
        con.close()
    if not row:
        raise SystemExit(f"No extractions found for schema '{schema_name}' in {db_path}")
    return row[0]


def _corpus_stats(df: pl.DataFrame) -> dict:
    """Compute high-level parquet statistics."""
    total = len(df)

    # European count
    n_european = 0
    if "is_european" in df.columns:
        n_european = int(df["is_european"].cast(pl.Boolean).fill_null(False).sum())

    # Papers by year
    papers_by_year: dict[str, int] = {}
    if "year" in df.columns:
        ydf = (
            df.filter(pl.col("year").is_not_null())
            .group_by("year")
            .agg(pl.len().alias("n"))
            .sort("year")
        )
        papers_by_year = {str(r["year"]): r["n"] for r in ydf.to_dicts()}

    # Top 10 journals
    top_journals: list[dict] = []
    if "journal" in df.columns:
        titles = [_journal_title(v) for v in df["journal"].to_list()]
        counts = Counter(t for t in titles if t)
        top_journals = [{"name": name, "count": cnt} for name, cnt in counts.most_common(10)]

    # Language breakdown
    languages: dict[str, int] = {}
    if "language" in df.columns:
        lang_df = (
            df.filter(pl.col("language").is_not_null())
            .group_by("language")
            .agg(pl.len().alias("n"))
            .sort("n", descending=True)
        )
        languages = {r["language"]: r["n"] for r in lang_df.to_dicts()}

    # Top funding agencies (from structured grants column — one count per paper per unique agency)
    top_funding_agencies: list[dict] = []
    if "grants" in df.columns:
        agency_counts: Counter = Counter()
        for raw in df["grants"].to_list():
            if not raw:
                continue
            try:
                grants_list = json.loads(raw)
                seen: set[str] = set()
                for g in grants_list:
                    if isinstance(g, dict) and g.get("agency"):
                        a = g["agency"].strip()
                        if a and a not in seen:
                            agency_counts[a] += 1
                            seen.add(a)
            except Exception:
                pass
        top_funding_agencies = [
            {"name": name, "count": cnt} for name, cnt in agency_counts.most_common(10)
        ]

    return {
        "total_papers": total,
        "european_papers": n_european,
        "european_pct": round(100 * n_european / total, 2) if total else 0,
        "papers_by_year": papers_by_year,
        "top_journals": top_journals,
        "languages": languages,
        "top_funding_agencies": top_funding_agencies,
    }


def _load_scope_lookup(output_dir: Path) -> dict[str, str]:
    """Load journal_name → scope from journal_registry.csv. Returns {} if missing."""
    registry_path = output_dir / "journal_registry.csv"
    if not registry_path.exists():
        return {}
    lookup: dict[str, str] = {}
    try:
        with registry_path.open(encoding="utf-8", newline="") as f:
            for row in csv.DictReader(f):
                name = (row.get("journal_name") or "").strip()
                scope = (row.get("scope") or "").strip()
                if name and scope:
                    lookup[name.lower()] = scope
    except Exception:
        return {}
    return lookup


def _journal_scope_stats(df: pl.DataFrame, output_dir: Path) -> dict[str, dict[str, int]]:
    """Compute paper counts per journal scope using journal_registry.csv annotations.

    Returns {scope: {total: N, european: N}} for scopes present in the registry.
    Returns {} if registry is missing or has no scope annotations yet.
    """
    scope_lookup = _load_scope_lookup(output_dir)
    if not scope_lookup:
        return {}

    read_cols = ["journal"]
    if "is_european" in df.columns:
        read_cols.append("is_european")

    distribution: dict[str, dict[str, int]] = {}
    for row in df.select(read_cols).to_dicts():
        title = _journal_title(row.get("journal"))
        if not title:
            continue
        scope = scope_lookup.get(title.lower(), "Other")
        if scope not in distribution:
            distribution[scope] = {"total": 0, "european": 0}
        distribution[scope]["total"] += 1
        if row.get("is_european"):
            distribution[scope]["european"] += 1

    return distribution


def _load_extractions(
    db_path: Path,
    schema_name: str,
    run_id: str,
    field_names: list[str],
    parquet_df: pl.DataFrame,
) -> tuple[dict, list[dict]]:
    """
    Returns (run_summary, extractions_list).
    extractions_list rows are joined with parquet for title/year/journal.
    """
    con = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
    con.row_factory = sqlite3.Row
    try:
        rows = con.execute(
            """
            SELECT pmid, extracted_json, raw_response, fix_applied, fix_history,
                   error_label, generation_id, model_used, provider_used,
                   cost_usd, prompt_tokens, completion_tokens, reasoning_tokens,
                   cached_tokens, latency_ms, created_at
            FROM extractions
            WHERE schema_name = ? AND run_id = ?
            ORDER BY created_at
            """,
            (schema_name, run_id),
        ).fetchall()
    finally:
        con.close()

    # Build parquet lookup: pmid → {title, year, journal_title, is_european}
    lookup_cols = ["pmid"]
    for col in ("title", "year", "journal", "is_european"):
        if col in parquet_df.columns:
            lookup_cols.append(col)
    lookup = {
        r["pmid"]: r for r in parquet_df.select(lookup_cols).to_dicts()
    }

    n_success = n_failed = n_repaired = 0
    total_cost = 0.0
    model_counter: Counter = Counter()
    field_distributions: dict[str, Counter] = {fn: Counter() for fn in field_names}

    extractions: list[dict] = []
    for row in rows:
        pmid = row["pmid"]
        extracted = _safe_json(row["extracted_json"])
        is_success = row["error_label"] is None and extracted is not None

        if is_success:
            n_success += 1
        else:
            n_failed += 1
        if row["fix_applied"]:
            n_repaired += 1
        if row["cost_usd"]:
            total_cost += row["cost_usd"]
        if row["model_used"]:
            model_counter[row["model_used"]] += 1

        # Merge field values into distributions
        field_values: dict[str, object] = {}
        if extracted and isinstance(extracted, dict):
            for fn in field_names:
                val = extracted.get(fn)
                if val is not None:
                    field_distributions[fn][str(val)] += 1
                    field_values[fn] = val

        paper = lookup.get(pmid, {})
        entry: dict = {
            "pmid": pmid,
            "title": paper.get("title"),
            "year": paper.get("year"),
            "journal_title": _journal_title(paper.get("journal")),
            "is_european": paper.get("is_european"),
            "model_used": row["model_used"],
            "provider_used": row["provider_used"],
            "cost_usd": row["cost_usd"],
            "fix_applied": row["fix_applied"],
            "error_label": row["error_label"],
            "latency_ms": row["latency_ms"],
            "created_at": row["created_at"],
            **field_values,
        }
        extractions.append(entry)

    run_summary = {
        "run_id": run_id,
        "schema_name": schema_name,
        "n_extracted": len(rows),
        "n_success": n_success,
        "n_failed": n_failed,
        "n_repaired": n_repaired,
        "total_cost_usd": round(total_cost, 6),
        "coverage_of_total_pct": round(100 * len(rows) / max(1, len(parquet_df)), 2),
        "models_used": dict(model_counter),
    }

    return run_summary, extractions, {k: dict(v) for k, v in field_distributions.items()}


# ── XLSX export ────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_TITLE_FONT  = Font(bold=True, size=12)
_BOLD_FONT   = Font(bold=True, size=10)

# Relevance colours
_REL_FILLS = {
    "Main":        PatternFill("solid", fgColor="C6EFCE"),
    "Secondary":   PatternFill("solid", fgColor="FFEB9C"),
    "NotRelevant": PatternFill("solid", fgColor="EDEDED"),
    "Irrelevant":  PatternFill("solid", fgColor="EDEDED"),
}
# Specialty colours
_SPEC_FILLS = {
    "Radiology":   PatternFill("solid", fgColor="BDD7EE"),
    "Cardiology":  PatternFill("solid", fgColor="FCE4D6"),
    "Unclear":     PatternFill("solid", fgColor="F2F2F2"),
    "NotReported": PatternFill("solid", fgColor="F2F2F2"),
}


def _header_row(ws, values: list, row: int = 1) -> None:
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _autofit(ws, min_w: int = 8, max_w: int = 60) -> None:
    for col_cells in ws.columns:
        length = max(
            len(str(c.value)) if c.value is not None else 0
            for c in col_cells
        )
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, min_w), max_w)


def save_xlsx(snapshot: dict, xlsx_path: Path) -> None:
    """Save a human-readable Excel workbook from a snapshot JSON.

    Sheets:
      1. Summary        — key metrics in a vertical key-value layout
      2. Distributions  — field-level classification breakdowns with percentages
      3. Top Statistics — top journals, funding agencies, papers by year
      4. Extractions    — one row per paper with all extracted fields (colour-coded)
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    cs = snapshot["corpus_stats"]
    rs = snapshot["run_summary"]
    sc = snapshot["search_config"]
    ec = snapshot["extraction_config"]
    fd = snapshot["field_distributions"]
    extractions = snapshot["extractions"]

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    ws.freeze_panes = "B1"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50

    rows_summary = [
        ("SEARCH", None),
        ("Project name",      sc.get("project_name", "")),
        ("Date range",        f"{sc.get('start_date', '')} → {sc.get('end_date', '')}"),
        ("Total papers",      cs.get("total_papers", 0)),
        ("European papers",   cs.get("european_papers", 0)),
        ("European %",        f"{cs.get('european_pct', 0):.1f}%"),
        ("", None),
        ("EXTRACTION", None),
        ("Schema name",       ec.get("schema_name", "")),
        ("Run ID",            rs.get("run_id", "")),
        ("Model",             ec.get("model", "")),
        ("Papers processed",  rs.get("n_extracted", 0)),
        ("Successfully extracted", rs.get("n_success", 0)),
        ("Failed",            rs.get("n_failed", 0)),
        ("Repaired by AI",    rs.get("n_repaired", 0)),
        ("Total cost (USD)",  f"${rs.get('total_cost_usd', 0):.4f}"),
        ("Coverage of corpus", f"{rs.get('coverage_of_total_pct', 0):.1f}%"),
        ("", None),
        ("PIPELINE", None),
        ("Steps run",         " → ".join(sc.get("steps", []))),
        ("Snapshot generated", snapshot.get("metadata", {}).get("generated_at", "")),
    ]

    for r, (key, val) in enumerate(rows_summary, 1):
        cell_k = ws.cell(row=r, column=1, value=key)
        if val is None:
            cell_k.font = _TITLE_FONT
            cell_k.fill = PatternFill("solid", fgColor="D9E1F2")
        else:
            cell_k.font = _BOLD_FONT
            ws.cell(row=r, column=2, value=val)

    # ── Sheet 2: Distributions ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Field Distributions")
    ws2.freeze_panes = "A2"
    _header_row(ws2, ["Field", "Value", "Count", "% of field total"], row=1)

    r = 2
    for field_name, dist in fd.items():
        total = max(sum(dist.values()), 1)
        for val, cnt in sorted(dist.items(), key=lambda x: -x[1]):
            ws2.cell(row=r, column=1, value=field_name)
            ws2.cell(row=r, column=2, value=val)
            ws2.cell(row=r, column=3, value=cnt)
            ws2.cell(row=r, column=4, value=round(100 * cnt / total, 1))
            r += 1

    _autofit(ws2)

    # ── Sheet 3: Top Statistics ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Top Statistics")
    ws3.freeze_panes = "A2"

    # Top journals
    ws3.cell(row=1, column=1, value="Top Journals").font = _TITLE_FONT
    _header_row(ws3, ["Journal", "Paper Count", "Scope"], row=2)
    for i, j in enumerate(cs.get("top_journals", []), 3):
        ws3.cell(row=i, column=1, value=j.get("name", ""))
        ws3.cell(row=i, column=2, value=j.get("count", 0))
        ws3.cell(row=i, column=3, value=j.get("scope", ""))

    # Top funding agencies (offset 2 columns to the right)
    offset_col = 5
    ws3.cell(row=1, column=offset_col, value="Top Funding Agencies").font = _TITLE_FONT
    _header_row(ws3, ["Agency", "Paper Count"], row=2)
    # Fix: use offset_col for these
    for col_i, hdr in enumerate(["Agency", "Paper Count"], offset_col):
        cell = ws3.cell(row=2, column=col_i, value=hdr)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for i, a in enumerate(cs.get("top_funding_agencies", []), 3):
        ws3.cell(row=i, column=offset_col, value=a.get("name", ""))
        ws3.cell(row=i, column=offset_col + 1, value=a.get("count", 0))

    # Papers by year
    yr_start_row = max(len(cs.get("top_journals", [])), len(cs.get("top_funding_agencies", []))) + 4
    ws3.cell(row=yr_start_row, column=1, value="Papers by Year").font = _TITLE_FONT
    _header_row(ws3, ["Year", "Papers"], row=yr_start_row + 1)
    for i, (yr, cnt) in enumerate(sorted(cs.get("papers_by_year", {}).items()), yr_start_row + 2):
        ws3.cell(row=i, column=1, value=int(yr))
        ws3.cell(row=i, column=2, value=cnt)

    _autofit(ws3)

    # ── Sheet 4: Extractions ──────────────────────────────────────────────────
    ws4 = wb.create_sheet("Extractions")
    ws4.freeze_panes = "A2"

    # Detect field columns from field_distributions keys
    field_cols = list(fd.keys())
    base_cols = ["PMID", "Title", "Year", "Journal", "European?"]
    ai_cols   = [f.replace("_", " ").title() for f in field_cols]
    meta_cols = ["Model Used", "Fix Applied", "Error", "Cost (USD)", "Created At"]
    all_col_names = base_cols + ai_cols + meta_cols
    _header_row(ws4, all_col_names, row=1)

    # Map from internal key to xlsx column index
    field_col_idx = {fc: base_cols.__len__() + i + 1 for i, fc in enumerate(field_cols)}

    for r_i, ext in enumerate(extractions, 2):
        ws4.cell(row=r_i, column=1, value=str(ext.get("pmid", "") or ""))
        ws4.cell(row=r_i, column=2, value=str(ext.get("title", "") or "")[:200])
        ws4.cell(row=r_i, column=3, value=ext.get("year"))
        ws4.cell(row=r_i, column=4, value=str(ext.get("journal_title", "") or "")[:80])
        ws4.cell(row=r_i, column=5, value="Yes" if ext.get("is_european") else "No")

        for fi, fc in enumerate(field_cols):
            col_i = len(base_cols) + fi + 1
            raw_val = ext.get(fc)
            val = str(raw_val) if raw_val is not None else ""
            cell = ws4.cell(row=r_i, column=col_i, value=val)
            # Colour relevance and specialty cells
            if fc == "relevance":
                fill = _REL_FILLS.get(val)
                if fill:
                    cell.fill = fill
            elif "specialty" in fc:
                fill = _SPEC_FILLS.get(val)
                if fill:
                    cell.fill = fill

        offset = len(base_cols) + len(field_cols)
        ws4.cell(row=r_i, column=offset + 1, value=str(ext.get("model_used", "") or ""))
        ws4.cell(row=r_i, column=offset + 2, value=str(ext.get("fix_applied", "") or ""))
        ws4.cell(row=r_i, column=offset + 3, value=str(ext.get("error_label", "") or ""))
        cost = ext.get("cost_usd")
        ws4.cell(row=r_i, column=offset + 4, value=round(cost, 6) if cost else None)
        ws4.cell(row=r_i, column=offset + 5, value=str(ext.get("created_at", "") or ""))

    _autofit(ws4, max_w=50)
    # Wider title column
    ws4.column_dimensions["B"].width = 50

    wb.save(xlsx_path)


# ── main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Export extraction run to JSON snapshot")
    parser.add_argument("--config", required=True, help="Path to project YAML config")
    parser.add_argument("--run-id", default=None, help="run_id to export (default: latest)")
    parser.add_argument("--out", default=None, help="Output JSON path (default: auto)")
    args = parser.parse_args()

    config_path = Path(args.config)
    if not config_path.exists():
        raise SystemExit(f"Config not found: {config_path}")

    cfg = _load_yaml(config_path)
    output_dir = Path(cfg.get("general", {}).get("output_dir", "output"))
    parquet_path = output_dir / "papers.parquet"
    db_path = output_dir / "extractions.db"

    extract_cfg = cfg.get("extract", {})
    fetch_cfg = cfg.get("fetch", {})
    schema_name = extract_cfg.get("schema_name", "")
    field_names = [f["name"] for f in extract_cfg.get("fields", [])]

    if not parquet_path.exists():
        raise SystemExit(f"papers.parquet not found at {parquet_path}")
    if not db_path.exists():
        raise SystemExit(f"extractions.db not found at {db_path} — run extract step first")

    print(f"Reading parquet: {parquet_path}")
    df = pl.read_parquet(parquet_path, memory_map=False)
    corpus_stats = _corpus_stats(df)
    corpus_stats["journal_scope"] = _journal_scope_stats(df, output_dir)

    # Enrich top_journals with scope from registry
    scope_lookup = _load_scope_lookup(output_dir)
    for j in corpus_stats["top_journals"]:
        j["scope"] = scope_lookup.get(j["name"].lower(), "Other")
    print(f"  {corpus_stats['total_papers']:,} papers  |  {corpus_stats['european_papers']:,} European ({corpus_stats['european_pct']}%)")
    if corpus_stats["journal_scope"]:
        for scope, counts in sorted(corpus_stats["journal_scope"].items(), key=lambda x: -x[1]["total"]):
            print(f"    {scope}: {counts['total']} total / {counts['european']} European")

    run_id = _pick_run_id(db_path, schema_name, args.run_id)
    print(f"Reading extractions: schema={schema_name!r}  run_id={run_id!r}")
    run_summary, extractions, field_distributions = _load_extractions(
        db_path, schema_name, run_id, field_names, df
    )
    print(f"  {run_summary['n_success']:,} success  |  {run_summary['n_failed']} failed  |  {run_summary['n_repaired']} repaired  |  ${run_summary['total_cost_usd']:.4f}")
    for fname, fdist in field_distributions.items():
        total_f = sum(fdist.values())
        print(f"    {fname}: {total_f} classified")

    snapshot = {
        "metadata": {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "generator": "scripts/per-project/export_run_json.py",
            "config_path": str(config_path),
        },
        "search_config": {
            "project_name": cfg.get("general", {}).get("project_name", ""),
            "query": fetch_cfg.get("query", "").strip(),
            "start_date": fetch_cfg.get("start_date", ""),
            "end_date": fetch_cfg.get("end_date", ""),
            "max_results": fetch_cfg.get("max_results", 0),
            "output_dir": str(output_dir),
            "steps": cfg.get("steps", []),
        },
        "extraction_config": {
            "schema_name": schema_name,
            "run_id": run_id,
            "model": extract_cfg.get("model", ""),
            "fallback_models": extract_cfg.get("fallback_models", []),
            "filter_column": extract_cfg.get("filter_column", ""),
            "user_instruction": extract_cfg.get("user_instruction", "").strip(),
            "fields": extract_cfg.get("fields", []),
            "max_cost_usd": extract_cfg.get("max_cost_usd", 0),
            "concurrency": extract_cfg.get("concurrency", 20),
        },
        "corpus_stats": corpus_stats,
        "run_summary": run_summary,
        "field_distributions": field_distributions,
        "extractions": extractions,
    }

    # Compute top_journals for original research (from extractions list)
    orig_journal_counts: Counter = Counter()
    for ext in extractions:
        st = str(ext.get("study_type", "") or "")
        # Accept both raw and normalised forms
        if st in ("OriginalResearch", "Original Research", "OriginalResearch"):
            jt = str(ext.get("journal_title", "") or "").strip()
            if jt:
                orig_journal_counts[jt] += 1
    corpus_stats["top_journals_orig"] = [
        {"name": name, "count": cnt, "scope": scope_lookup.get(name.lower(), "Other")}
        for name, cnt in orig_journal_counts.most_common(10)
    ]

    out_path = Path(args.out) if args.out else (
        output_dir / f"snapshot_{schema_name}_{run_id}.json"
    )
    out_path.write_text(json.dumps(snapshot, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
    print(f"\nSnapshot saved -> {out_path}  ({out_path.stat().st_size / 1024:.1f} KB)")

    xlsx_path = out_path.with_suffix(".xlsx")
    try:
        save_xlsx(snapshot, xlsx_path)
        print(f"Excel saved    -> {xlsx_path}  ({xlsx_path.stat().st_size / 1024:.0f} KB)")
    except Exception as exc:
        print(f"[warn] Excel export failed: {exc}")


if __name__ == "__main__":
    main()
